from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------- Styles ----------
HDR_FILL = PatternFill("solid", start_color="1F3864")
SUB_FILL = PatternFill("solid", start_color="2E75B6")
BAND_FILL = PatternFill("solid", start_color="D9E1F2")
WIN_FILL = PatternFill("solid", start_color="C6EFCE")
white_bold = Font(name="Arial", bold=True, color="FFFFFF", size=11)
white_bold_sm = Font(name="Arial", bold=True, color="FFFFFF", size=10)
bold = Font(name="Arial", bold=True, size=11)
blue = Font(name="Arial", color="0000FF", size=11)   # hardcoded inputs
black = Font(name="Arial", color="000000", size=11)  # formulas
note = Font(name="Arial", italic=True, color="595959", size=9)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ================= SHEET 1: SCORECARD =================
ws = wb.active
ws.title = "Scorecard"

ws["A1"] = "Trip & Events App Portfolio — Prioritization Scorecard"
ws["A1"].font = Font(name="Arial", bold=True, size=15, color="1F3864")
ws.merge_cells("A1:R1")
ws["A2"] = "Build order driven by Time-to-Cash + Strategic Fit (priority 1), then Market Size. Frameworks: weighted scorecard + RICE. Scores 1-5 (blue = your editable inputs)."
ws["A2"].font = note
ws.merge_cells("A2:R2")

# Weights block
ws["A4"] = "WEIGHTS (priority-tuned)"
ws["A4"].font = bold
wlabels = ["Build Ease", "Time-to-Cash", "Market Size", "Monetization", "Validation Ease", "Strategic Fit"]
wvals =   [1.5,          3,              2,             1.5,            1,                 3]
for i, (lab, val) in enumerate(zip(wlabels, wvals)):
    c = ws.cell(row=5, column=1+i, value=lab); c.font = white_bold_sm; c.fill = SUB_FILL; c.alignment = ctr; c.border = border
    c2 = ws.cell(row=6, column=1+i, value=val); c2.font = blue; c2.alignment = ctr; c2.border = border
ws["H5"] = "Weight sum"; ws["H5"].font = white_bold_sm; ws["H5"].fill = SUB_FILL; ws["H5"].alignment = ctr; ws["H5"].border = border
ws["H6"] = "=SUM(A6:F6)"; ws["H6"].font = black; ws["H6"].alignment = ctr; ws["H6"].border = border

# Main table header
hdr_row = 9
headers = ["#", "Product (working name)", "Segment", "One-liner",
           "Build\nEase", "Time-to\nCash", "Market\nSize", "Monet-\nization", "Valid'n\nEase", "Strategic\nFit",
           "Weighted\nScore /100", "Rank",
           "RICE\nReach", "RICE\nImpact", "RICE\nConf.", "RICE\nEffort(pm)", "RICE\nScore"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=hdr_row, column=j, value=h); c.font = white_bold_sm; c.fill = HDR_FILL; c.alignment = ctr; c.border = border

# Data: [#, name, segment, one-liner, ease, ttc, market, monet, valid, fit, reach, impact, conf, effort]
rows = [
 [1,"SplitTrip","Group finance","Trip/event-framed bill splitting with invite-to-split viral loop",5,4,5,3,5,5, 8,2,0.8,1.5],
 [2,"EventList","Events / social","Guest list + RSVP + countdown for parties, dinners, weddings",5,3,4,3,5,3, 6,1,0.7,1.5],
 [3,"TripBoard","Travel planning","Real-time collaborative itinerary + shared packing/checklists",4,3,4,3,4,4, 5,1,0.6,2.5],
 [4,"TripMap","Location / GIS","Opt-in group live location + pin & tag special places on a map",2,2,3,2,2,4, 3,1,0.4,4.0],
 [5,"TripReel","Memories / media","Photos & videos on a trip map + auto-generated recap video",2,3,4,4,2,5, 5,2,0.6,4.0],
 [6,"The Suite","Integrated platform","All-in-one trip super-app: bills + itinerary + map + memories",1,2,5,5,1,5, 8,3,0.4,9.0],
]

r0 = hdr_row + 1
for i, d in enumerate(rows):
    r = r0 + i
    vals = [d[0], d[1], d[2], d[3]]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = bold if j in (1,2) else black
        c.alignment = ctr if j==1 else left
        c.border = border
        if i % 2 == 1: c.fill = BAND_FILL
    # score inputs (blue) cols 5-10
    for j, v in enumerate(d[4:10], start=5):
        c = ws.cell(row=r, column=j, value=v); c.font = blue; c.alignment = ctr; c.border = border
        if i % 2 == 1: c.fill = BAND_FILL
    # weighted score col 11
    c = ws.cell(row=r, column=11,
        value=f"=SUMPRODUCT(E{r}:J{r},$A$6:$F$6)/(5*$H$6)*100")
    c.font = Font(name="Arial", bold=True, size=11); c.alignment = ctr; c.border = border; c.number_format='0.0'
    if i % 2 == 1: c.fill = BAND_FILL
    # rank col 12
    c = ws.cell(row=r, column=12, value=f"=RANK(K{r},$K${r0}:$K${r0+len(rows)-1})")
    c.font = black; c.alignment = ctr; c.border = border
    if i % 2 == 1: c.fill = BAND_FILL
    # RICE inputs cols 13-16
    for j, v in enumerate(d[10:14], start=13):
        c = ws.cell(row=r, column=j, value=v); c.font = blue; c.alignment = ctr; c.border = border
        if j==15: c.number_format='0%'
        if i % 2 == 1: c.fill = BAND_FILL
    # RICE score col 17
    c = ws.cell(row=r, column=17, value=f"=M{r}*N{r}*O{r}/P{r}")
    c.font = Font(name="Arial", bold=True, size=11); c.alignment = ctr; c.border = border; c.number_format='0.00'
    if i % 2 == 1: c.fill = BAND_FILL

# Legend / scale notes
lr = r0 + len(rows) + 2
ws.cell(row=lr, column=1, value="SCORING SCALE (1 = poor/hard, 5 = excellent/easy)").font = bold
scale = [
 ("Build Ease","5 = ship in days (CRUD/forms); 1 = heavy infra (real-time GPS, video pipeline)"),
 ("Time-to-Cash","5 = can charge at launch; 1 = long runway before willingness to pay"),
 ("Market Size","5 = mass-market everyone; 1 = narrow niche"),
 ("Monetization","5 = strong proven willingness to pay; 1 = users expect free"),
 ("Validation Ease","5 = cheap to test real demand; 1 = expensive to even prove"),
 ("Strategic Fit","5 = core to the trip-memory endgame; 1 = adjacent / off-path"),
 ("RICE","Score = Reach x Impact x Confidence / Effort(person-months). Reach in relative k-users/2 quarters; Impact 0.25-3; Conf as %."),
]
for i,(k,v) in enumerate(scale):
    ws.cell(row=lr+1+i, column=1, value=k).font = bold
    cc = ws.cell(row=lr+1+i, column=2, value=v); cc.font = note; ws.merge_cells(start_row=lr+1+i,start_column=2,end_row=lr+1+i,end_column=10)

# column widths
widths = [4,18,17,40,8,8,8,8,8,9,11,6,7,7,7,9,8]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hdr_row].height = 34
for i in range(len(rows)):
    ws.row_dimensions[r0+i].height = 30
ws.freeze_panes = "B10"

# ================= SHEET 2: SEQUENCING =================
ws2 = wb.create_sheet("Build Sequence")
ws2["A1"] = "Recommended Ship → Cash → Learn Sequence"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")
ws2.merge_cells("A1:F1")
seq_hdr = ["Wave","Product","Ship target","Monetization at this wave","Kill / Go signal (what real pulse tells you)","Funds the next"]
for j,h in enumerate(seq_hdr, start=1):
    c = ws2.cell(row=3, column=j, value=h); c.font = white_bold_sm; c.fill = HDR_FILL; c.alignment = ctr; c.border = border
seq = [
 ["Wave 1","SplitTrip","Weeks 1-4","Freemium: free core; $3-5/mo for receipt scan, multi-currency, unlimited trips","Go if invited friends activate & free->paid > ~3-5%. Crowded vs Splitwise, so watch retention.","Cash + a warm user base already grouped into trips"],
 ["Wave 2","EventList","Weeks 5-8","Mostly free (Partiful owns free); charge only on weddings / premium event pages","Go if events drive new groups in cheaply. If CAC poor, keep as a free feeder, not a product.","More groups; widens top of funnel"],
 ["Wave 2b","TripBoard","Weeks 7-12","$30-40/yr Pro (matches Wanderlog) once collab itinerary is sticky","Go if daily-active during trips rises. Wanderlog is strong — only push if engagement is real.","Daily engagement = the audience for memories"],
 ["Wave 3","TripMap","Months 4-6","Bundled into Pro; not sold standalone (Life360/Find My own this)","Build only after Wave 1-2 prove people stick. Privacy opt-in is mandatory.","Captures the raw route + places the recap needs"],
 ["Wave 4","TripReel","Months 6-9","One-time recap video unlock and/or printed book (Polarsteps model, €36+)","The emotional + viral payoff. Recap shares = your cheapest acquisition.","Highest-margin sales + organic growth engine"],
 ["Wave 5","The Suite","Months 9+","Single Pro bundle across all modules; multiple revenue streams","Only assemble once each module has proven demand independently.","The endgame — defensible because no incumbent owns the whole loop"],
]
for i,row in enumerate(seq):
    r = 4+i
    for j,v in enumerate(row, start=1):
        c = ws2.cell(row=r, column=j, value=v); c.border = border
        c.alignment = ctr if j in (1,3) else left
        c.font = bold if j in (1,2) else black
        if i % 2 == 1: c.fill = BAND_FILL
ws2_w = [9,14,12,34,40,34]
for i,w in enumerate(ws2_w, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for i in range(len(seq)):
    ws2.row_dimensions[4+i].height = 56
ws2.row_dimensions[3].height = 30

wb.save("/sessions/tender-quirky-bardeen/mnt/outputs/Trip_App_Prioritization_Scorecard.xlsx")
print("saved")
